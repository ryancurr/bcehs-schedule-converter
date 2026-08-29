import datetime as dt
import re
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange


# ===================== SHARED RULES =====================

STUDENT_ALIASES = {
    "rory": "Rory-lynn Bradshaw",
}

EXCLUDE_STUDENTS = {
    "jadyn",
    "jadyn langley",
}

PARTNER_MARKERS = {
    "partner", "parnter", "partnre", "parter", "prtnr",
}

# Handles one or more repeated suffixes, e.g.
# "Christopher Ramirez - Columbia - Columbia"
COLUMBIA_SUFFIX_PAT = re.compile(
    r"(?:\s*-\s*Columbia\s*)+$",
    re.I,
)

# Header examples supported:
# Sep/3, Sept/3, Sep 3, Sept 3, September 3
HEADER_DATE_PAT = re.compile(
    r"\b([A-Za-z]{3,9})\s*(?:/|\s+|-)\s*(\d{1,2})\b",
    re.I,
)

TIME_RANGE_PAT = re.compile(
    r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})|"
    r"(\d{3,4})\s*[-–]\s*(\d{3,4})"
)

CODE_TOKEN_PAT = re.compile(r"^\d{3}[A-Za-z0-9]+$")
AMBULANCE_PAT = re.compile(r"^(\d{3}[A-Z]\d)", re.I)

MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

OLE_XLS_MAGIC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"


def bytes_to_filelike(b: bytes):
    import io
    return io.BytesIO(b)


def is_legacy_xls(file_bytes: bytes) -> bool:
    return file_bytes[:8] == OLE_XLS_MAGIC


# ===================== MERGED CELLS: OPENPYXL =====================


def build_merged_map(ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    For .xlsx/.xlsm files, map every cell inside a merged range to
    that range's top-left cell. Coordinates are 1-based.
    """
    merged_map: Dict[Tuple[int, int], Tuple[int, int]] = {}

    for rng in ws.merged_cells.ranges:
        cr = CellRange(str(rng))
        top_left = (cr.min_row, cr.min_col)

        for r in range(cr.min_row, cr.max_row + 1):
            for c in range(cr.min_col, cr.max_col + 1):
                merged_map[(r, c)] = top_left

    return merged_map


def get_cell_value(
    ws,
    r: int,
    c: int,
    merged_map: Dict[Tuple[int, int], Tuple[int, int]],
):
    if r < 1 or c < 1:
        return None

    value = ws.cell(r, c).value
    if value is not None:
        return value

    top_left = merged_map.get((r, c))
    if top_left:
        return ws.cell(top_left[0], top_left[1]).value

    return None


# ===================== SHIFT PARSING =====================


def norm_hhmm(x: str) -> str:
    x = str(x)

    if len(x) == 3:
        hour = int(x[0])
        minute = int(x[1:])
    else:
        hour = int(x[:-2])
        minute = int(x[-2:])

    return f"{hour:02d}:{minute:02d}"


def parse_shift(text: str) -> Optional[dict]:
    if not isinstance(text, str):
        return None

    t = text.replace("\n", " ").strip()
    if not t or t == "\\":
        return None

    code = ""
    for token in t.split():
        if CODE_TOKEN_PAT.match(token):
            code = token
            break

    start = ""
    end = ""

    match = TIME_RANGE_PAT.search(t)
    if match:
        if match.group(1):
            start = match.group(1)
            end = match.group(2)
        else:
            start = norm_hhmm(match.group(3))
            end = norm_hhmm(match.group(4))

    if not (code or start or end):
        return None

    station = code[:3] if code else ""

    ambulance = ""
    if code:
        ambulance_match = AMBULANCE_PAT.match(code)
        if ambulance_match:
            ambulance = ambulance_match.group(1)

    return {
        "raw": t,
        "code": code,
        "start": start,
        "end": end,
        "station": station,
        "ambulance": ambulance,
    }


# ===================== NAMES =====================


def format_preceptor_one(name: str) -> str:
    s = re.sub(r"\s+", " ", str(name).strip())

    if "," in s:
        last, rest = s.split(",", 1)
        return re.sub(
            r"\s+",
            " ",
            f"{rest.strip()} {last.strip()}".strip(),
        )

    return s


def format_preceptor(name: str) -> str:
    """
    ACP can contain multiple preceptors separated by '/'.
    Example: "Wilson, Travis / Johnston, Heather".
    """
    if not isinstance(name, str):
        return ""

    parts = [p.strip() for p in name.split("/") if p.strip()]
    parts = [format_preceptor_one(p) for p in parts]

    return " / ".join(parts)


def is_partner_marker(s: str) -> bool:
    if not isinstance(s, str):
        return False

    t = s.strip().lower()

    if t in PARTNER_MARKERS:
        return True

    return ("partner" in t) or ("parnter" in t)


def clean_student_name(raw: str) -> str:
    """
    Shared name cleanup after eligibility has been determined.
    """
    if not isinstance(raw, str):
        return ""

    s = re.sub(r"\s+", " ", raw.strip())
    if not s:
        return ""

    if is_partner_marker(s):
        return ""

    if s.lower() in {"student", "n/a", "na", "tbd"}:
        return ""

    # Strip one or more Columbia suffixes.
    s = re.sub(COLUMBIA_SUFFIX_PAT, "", s).strip()

    if s.lower() in STUDENT_ALIASES:
        s = STUDENT_ALIASES[s.lower()]

    if s.lower() in EXCLUDE_STUDENTS:
        return ""

    return s


def normalize_student_pcp(raw: str) -> str:
    """
    PCP rule:
    BCEHS does not consistently add '- Columbia', so PCP does NOT
    require that suffix. If present, it is removed.
    """
    return clean_student_name(raw)


def normalize_student_acp(raw: str) -> str:
    """
    ACP rule:
    Only rows explicitly marked '- Columbia' are Columbia students.
    This prevents PARTNER, TBD, Sar Tech, and other BCEHS entries from
    being imported as students.
    """
    if not isinstance(raw, str):
        return ""

    s = re.sub(r"\s+", " ", raw.strip())
    if not s:
        return ""

    if not COLUMBIA_SUFFIX_PAT.search(s):
        return ""

    return clean_student_name(s)


# ===================== DATE HEADERS =====================


def parse_header_date(value, year: int) -> Optional[dt.date]:
    if not isinstance(value, str):
        return None

    match = HEADER_DATE_PAT.search(value)
    if not match:
        return None

    # "Sept" and "September" both normalize to "sep".
    month_key = match.group(1).lower()[:3]
    month = MONTHS.get(month_key)
    if not month:
        return None

    day = int(match.group(2))

    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def parse_header_dates(ws, year: int, start_col: int) -> Dict[int, dt.date]:
    """
    OpenPyXL version. Coordinates are 1-based.
    """
    col_dates: Dict[int, dt.date] = {}

    for c in range(start_col, ws.max_column + 1):
        parsed = parse_header_date(ws.cell(1, c).value, year)
        if parsed:
            col_dates[c] = parsed

    return col_dates


# ===================== GROUP HEADERS =====================


def is_group_header(text: str) -> bool:
    """
    Existing group-header detection used by PCP.
    """
    if not isinstance(text, str):
        return False

    s = text.strip()

    if s.upper().startswith("STUDENT"):
        return False

    if "," in s:
        return False

    if s.upper() == s and len(s) >= 3:
        return True

    if any(
        key in s
        for key in [
            "Metro",
            "Vancouver",
            "Fraser",
            "Interior",
            "Island",
            "&",
            "SEA TO SKY",
            "COASTAL",
        ]
    ):
        return len(s.split()) >= 2

    return False


def is_acp_group_header(text: str) -> bool:
    """
    ACP now uses section labels such as:
      C - Surrey (249, 253), Delta (266), White Rock (254), Langley (267)

    Keep support for older ACP group-heading styles as well.
    """
    if not isinstance(text, str):
        return False

    s = text.strip()
    if not s or s.upper().startswith("STUDENT"):
        return False

    if re.match(r"^[A-Z]\s*-\s*.+\(\s*\d{3}", s, re.I):
        return True

    return is_group_header(s)


# ===================== PCP =====================


def extract_pcp_rows(wb, year: int) -> pd.DataFrame:
    """
    PCP (.xlsx/.xlsm):
    - multiple region sheets
    - dates begin at column B
    - student marker row is directly above preceptor/shift row
    - student names may be merged across several dates

    This pathway is intentionally kept separate from the ACP changes.
    """
    all_rows: List[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_map = build_merged_map(ws)
        col_dates = parse_header_dates(ws, year, start_col=2)

        if not col_dates:
            continue

        current_group = None

        for r in range(2, ws.max_row + 1):
            a = get_cell_value(ws, r, 1, merged_map)

            if not isinstance(a, str):
                continue

            a_str = a.strip()
            if not a_str:
                continue

            if a_str == "Preceptor" or a_str.upper().startswith("STUDENT"):
                continue

            if is_group_header(a_str):
                current_group = a_str
                continue

            preceptor = format_preceptor_one(a_str)

            student_marker = get_cell_value(ws, r - 1, 1, merged_map)
            if not (
                isinstance(student_marker, str)
                and student_marker.strip().upper().startswith("STUDENT")
            ):
                continue

            for c, date in col_dates.items():
                shift_value = get_cell_value(ws, r, c, merged_map)
                shift = parse_shift(shift_value) if isinstance(shift_value, str) else None

                if not shift:
                    continue

                student_raw = get_cell_value(ws, r - 1, c, merged_map)
                student = normalize_student_pcp(
                    student_raw if isinstance(student_raw, str) else ""
                )

                if not student:
                    continue

                location = current_group if current_group else sheet_name
                if current_group and current_group != sheet_name:
                    location = f"{sheet_name} - {current_group}"

                all_rows.append(
                    {
                        "Student Name": student,
                        "Date (YYYY-MM-DD)": date.isoformat(),
                        "Start Time (HH:MM)": shift["start"],
                        "End Time (HH:MM)": shift["end"],
                        "Location": location,
                        "Station": shift["station"],
                        "Ambulance Number": shift["ambulance"],
                        "Preceptor": preceptor,
                    }
                )

    return pd.DataFrame(all_rows)


# ===================== GENERIC ACP GRID =====================


def is_student_marker(value) -> bool:
    return (
        isinstance(value, str)
        and value.strip().upper().startswith("STUDENT")
    )


def extract_acp_rows_from_grid(
    max_row: int,
    max_col: int,
    get_value: Callable[[int, int], object],
    year: int,
) -> pd.DataFrame:
    """
    Generic ACP extractor used by BOTH:
      - modern .xlsx/.xlsm files
      - legacy binary .xls files

    get_value uses zero-based row/column coordinates and must already
    resolve merged-cell values.
    """
    col_dates: Dict[int, dt.date] = {}

    # Do not assume where the first date column is. The September 2026
    # file added "Mentor / Preceptor" in column C, moving dates to D.
    for c in range(max_col):
        parsed = parse_header_date(get_value(0, c), year)
        if parsed:
            col_dates[c] = parsed

    if not col_dates:
        return pd.DataFrame()

    rows: List[dict] = []
    pending_student_rows: List[int] = []

    for r in range(1, max_row):
        a = get_value(r, 0)

        if not isinstance(a, str) or not a.strip():
            continue

        a_str = a.strip()

        if a_str == "Preceptor":
            continue

        if is_acp_group_header(a_str):
            pending_student_rows = []
            continue

        if is_student_marker(a_str):
            pending_student_rows.append(r)
            continue

        # Any non-marker row here is the next preceptor row below the
        # accumulated STUDENT / STUDENT 1 / STUDENT 2 rows.
        preceptor = format_preceptor(a_str)
        student_rows = pending_student_rows
        pending_student_rows = []

        if not student_rows:
            continue

        for c, date in col_dates.items():
            shift_value = get_value(r, c)
            shift = parse_shift(shift_value) if isinstance(shift_value, str) else None

            if not shift:
                continue

            collected_students: List[str] = []

            for student_row in student_rows:
                raw = get_value(student_row, c)
                student = normalize_student_acp(
                    raw if isinstance(raw, str) else ""
                )

                if student:
                    collected_students.append(student)

            # De-duplicate while preserving source order.
            seen = set()
            students: List[str] = []

            for student in collected_students:
                key = student.lower()
                if key not in seen:
                    seen.add(key)
                    students.append(student)

            for student in students:
                rows.append(
                    {
                        "Student Name": student,
                        "Date (YYYY-MM-DD)": date.isoformat(),
                        "Start Time (HH:MM)": shift["start"],
                        "End Time (HH:MM)": shift["end"],
                        "Location": "ACP",
                        "Station": shift["station"],
                        "Ambulance Number": shift["ambulance"],
                        "Preceptor": preceptor,
                    }
                )

    return pd.DataFrame(rows)


# ===================== ACP: XLSX / XLSM =====================


def extract_acp_rows_openpyxl(wb, year: int) -> pd.DataFrame:
    ws = wb[wb.sheetnames[0]]
    merged_map = build_merged_map(ws)

    def get_value(r: int, c: int):
        # Generic ACP grid is zero-based; OpenPyXL is one-based.
        return get_cell_value(ws, r + 1, c + 1, merged_map)

    return extract_acp_rows_from_grid(
        max_row=ws.max_row,
        max_col=ws.max_column,
        get_value=get_value,
        year=year,
    )


# ===================== ACP: LEGACY .XLS =====================


def extract_acp_rows_xls(file_bytes: bytes, year: int) -> pd.DataFrame:
    """
    Read old binary Excel .xls files using xlrd.

    formatting_info=True is important because it exposes merged ranges,
    allowing multi-day student blocks to be expanded correctly.
    """
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "This ACP schedule is an old .xls file. Add xlrd==2.0.1 "
            "to requirements.txt so Streamlit can read it."
        ) from exc

    book = xlrd.open_workbook(
        file_contents=file_bytes,
        formatting_info=True,
    )

    sheet = book.sheet_by_index(0)

    merged_map: Dict[Tuple[int, int], Tuple[int, int]] = {}

    # xlrd merged ranges are (row_low, row_high, col_low, col_high)
    # with the high values EXCLUSIVE.
    for row_low, row_high, col_low, col_high in sheet.merged_cells:
        top_left = (row_low, col_low)

        for r in range(row_low, row_high):
            for c in range(col_low, col_high):
                merged_map[(r, c)] = top_left

    def raw_value(r: int, c: int):
        if r < 0 or c < 0 or r >= sheet.nrows or c >= sheet.ncols:
            return None

        return sheet.cell_value(r, c)

    def get_value(r: int, c: int):
        value = raw_value(r, c)

        # xlrd usually represents an empty cell as "" rather than None.
        if value not in (None, ""):
            return value

        top_left = merged_map.get((r, c))
        if top_left:
            return raw_value(top_left[0], top_left[1])

        return value

    return extract_acp_rows_from_grid(
        max_row=sheet.nrows,
        max_col=sheet.ncols,
        get_value=get_value,
        year=year,
    )


# ===================== PUBLIC API =====================


def extract_rows_from_workbook(
    workbook_bytes: bytes,
    year: int,
    mode: str,
) -> pd.DataFrame:
    mode = (mode or "").strip().upper()

    if mode == "ACP" and is_legacy_xls(workbook_bytes):
        return extract_acp_rows_xls(workbook_bytes, year)

    if mode == "PCP" and is_legacy_xls(workbook_bytes):
        raise ValueError(
            "PCP legacy .xls files are not supported. "
            "The current PCP schedule should remain .xlsx/.xlsm."
        )

    wb = load_workbook(
        filename=bytes_to_filelike(workbook_bytes),
        data_only=True,
    )

    if mode == "ACP":
        return extract_acp_rows_openpyxl(wb, year)

    return extract_pcp_rows(wb, year)


def apply_template_columns(
    extracted: pd.DataFrame,
    template_csv_path: str,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    template = pd.read_csv(template_csv_path)
    template_columns = list(template.columns)

    if extracted.empty:
        output = pd.DataFrame(columns=template_columns)
    else:
        output = extracted[template_columns].copy()

    debug = extracted.copy()

    return output, debug
